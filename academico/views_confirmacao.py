"""
Views para as features de Confirmação de Matrícula, Congelamento e Exportação.
Adiciona este ficheiro como academico/views_confirmacao.py
"""
import io
import json
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.template.loader import get_template

from xhtml2pdf import pisa

from .models import Aluno, ConfirmacaoMatricula, Nota, Presenca, Inscricao
from .services import get_ano_letivo_atual
from escola.models import Curso, Turma, Disciplina, Atribuicao, Professor
from .utils import _enviar_email_async


# ── Helpers de email ──────────────────────────────────────────────────────

def _email_confirmacao(confirmacao, tipo):
    """Envia emails automáticos conforme o status da confirmação de matrícula."""
    from django.conf import settings

    nome = (confirmacao.nome_completo or
            (confirmacao.aluno.usuario.get_full_name() if confirmacao.aluno else None) or
            (confirmacao.aluno.usuario.username if confirmacao.aluno else 'Aluno'))
    email = confirmacao.email

    if not email:
        return

    templates = {
        'EM_REVISAO': {
            'subject': '📋 Confirmação de Matrícula Recebida — Colégio Tarimba',
            'body': f"""
                <p>Olá, <strong>{nome}</strong>,</p>
                <p>A tua confirmação de matrícula para o ano lectivo <strong>{confirmacao.ano_letivo}</strong> foi recebida.</p>
                <p>Estamos a verificar os teus documentos e a disponibilidade de vagas na turma. Entraremos em contacto brevemente.</p>
            """,
        },
        'AGUARDANDO_PAGAMENTO': {
            'subject': '✅ Documentos Aprovados — Efectua o Pagamento',
            'body': f"""
                <p>Olá, <strong>{nome}</strong>,</p>
                <p>Os teus documentos foram verificados e aprovados.</p>
                <div style="background:#FFF3ED;border:1px solid #E8621A;border-radius:8px;padding:15px;margin:20px 0">
                    <p><strong>Próximo passo:</strong> Dirija-se ao Colégio Tarimba para efectuar o pagamento da taxa de confirmação de matrícula.</p>
                    <p style="font-size:13px;color:#666">Apresenta este email ou o teu número de processo.</p>
                </div>
            """,
        },
        'ATIVO': {
            'subject': '🎓 Matrícula Confirmada — Acesso Desbloqueado',
            'body': f"""
                <p>Olá, <strong>{nome}</strong>,</p>
                <p>O teu pagamento foi validado e a tua matrícula para <strong>{confirmacao.ano_letivo}</strong> está confirmada!</p>
                <p>O teu acesso ao Portal do Aluno foi desbloqueado. Podes aceder com as tuas credenciais habituais.</p>
                <div style="background:#f0fdf4;border:1px solid #22c55e;border-radius:8px;padding:15px;margin:20px 0">
                    <p><strong>Nova classe:</strong> {confirmacao.get_classe_nova_display()}</p>
                    <p><strong>Curso:</strong> {confirmacao.curso_novo or 'Mantido'}</p>
                </div>
            """,
        },
        'REJEITADO': {
            'subject': '❌ Confirmação de Matrícula Não Aprovada',
            'body': f"""
                <p>Olá, <strong>{nome}</strong>,</p>
                <p>Após análise, não foi possível aprovar a tua confirmação de matrícula.</p>
                {'<div style="background:#fff3f3;border-left:4px solid #e53e3e;padding:10px;margin:15px 0"><p><strong>Motivo:</strong> ' + (confirmacao.motivo_rejeicao or '') + '</p></div>' if confirmacao.motivo_rejeicao else ''}
                <p>Para mais informações, contacta a secretaria do colégio.</p>
            """,
        },
    }

    if tipo not in templates:
        return

    t = templates[tipo]
    html = f"""
    <html><body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:20px;">
      <div style="max-width:560px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1)">
        <div style="background:#E8621A;padding:20px;text-align:center;color:#fff">
          <h2 style="margin:0">Colégio Tarimba</h2>
          <p style="margin:5px 0 0;font-size:13px">Confirmação de Matrícula {confirmacao.ano_letivo}</p>
        </div>
        <div style="padding:30px">{t['body']}</div>
        <div style="background:#f9f9f9;padding:15px;text-align:center;font-size:11px;color:#999;border-top:1px solid #eee">
          Sistema Académico Tarimba · C.E.P. J.M.F.D - Calumbo
        </div>
      </div>
    </body></html>
    """

    _enviar_email_async(
        subject=t['subject'],
        message=f"Olá {nome}, o estado da tua confirmação de matrícula foi actualizado.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[email],
        html_message=html,
    )


# ── Feature 1: Confirmação de Matrícula ──────────────────────────────────

def confirmacao_matricula(request):
    """
    Formulário de confirmação de matrícula para alunos existentes ou visitantes.
    - Se aluno logado: preenche dados automaticamente
    - Se visitante: permite preencher todos os dados
    Calcula automaticamente a próxima classe para alunos já no sistema.
    """
    ano_letivo = get_ano_letivo_atual()
    aluno = None
    classe_atual = None
    proxima_classe = None
    precisa_escolher_curso = False
    cursos_ii_ciclo = None
    usuario_logado = request.user.is_authenticated

    def _achar_aluno_por_bi_e_nome(bi, nome):
        if not bi:
            return None
        aluno_obj = Aluno.objects.select_related('turma__curso', 'usuario').filter(usuario__bi_numero=bi).first()
        if aluno_obj:
            return aluno_obj
        if nome:
            partes = nome.strip().split()
            if len(partes) >= 2:
                first_name = partes[0]
                last_name = ' '.join(partes[1:])
                aluno_obj = Aluno.objects.select_related('turma__curso', 'usuario').filter(
                    usuario__first_name__iexact=first_name,
                    usuario__last_name__iexact=last_name
                ).first()
                if aluno_obj:
                    return aluno_obj
            return Aluno.objects.select_related('turma__curso', 'usuario').filter(
                usuario__first_name__iexact=nome.strip()
            ).first()
        return None

    def _ano_letivo_para_confirmacao(aluno_obj):
        if aluno_obj and aluno_obj.turma:
            return aluno_obj.turma.ano_letivo + 1
        return get_ano_letivo_atual()

    # Se logado e é aluno, obtém dados da turma
    if usuario_logado:
        try:
            aluno = Aluno.objects.select_related('turma__curso').get(usuario=request.user)
            ano_letivo = _ano_letivo_para_confirmacao(aluno)
            classe_atual = aluno.turma.classe if aluno.turma else None
            proxima_classe = ConfirmacaoMatricula.calcular_proxima_classe(classe_atual)
            precisa_escolher_curso = (str(classe_atual) == '9')
            cursos_ii_ciclo = Curso.objects.filter(
                tipo__in=['GERAL', 'TECNICO']
            ) if precisa_escolher_curso else None

            # Verifica se já submeteu para o ano de confirmação correto
            confirmacao_existente = ConfirmacaoMatricula.objects.filter(
                aluno=aluno, ano_letivo=ano_letivo
            ).first()
            if confirmacao_existente:
                return render(request, 'academico/confirmacao_estado.html', {
                    'confirmacao': confirmacao_existente,
                })
        except Aluno.DoesNotExist:
            # Usuário logado mas sem perfil de aluno
            # Permite preencher como visitante
            aluno = None

    classes_disponiveis = Turma.CLASSES_CHOICES
    form_data = request.POST if request.method == 'POST' else {}

    if request.method == 'POST':
        foto = request.FILES.get('foto_rosto')
        bi = request.POST.get('bi_numero', '').strip()
        email = request.POST.get('email', '').strip()
        nome_visitante = request.POST.get('nome_completo', '').strip() if not aluno else None
        classe_atual_post = request.POST.get('classe_atual') if not aluno else None
        curso_id = request.POST.get('curso_novo')

        # Se o visitante já corresponder a um aluno pelo BI ou nome, ligamos o registo
        if not aluno and bi:
            aluno = _achar_aluno_por_bi_e_nome(bi, nome_visitante)
            if aluno:
                ano_letivo = _ano_letivo_para_confirmacao(aluno)
            classe_atual = classe_atual_post or None
            proxima_classe = ConfirmacaoMatricula.calcular_proxima_classe(classe_atual)
            precisa_escolher_curso = (str(classe_atual) == '9')
            cursos_ii_ciclo = Curso.objects.filter(
                tipo__in=['GERAL', 'TECNICO']
            ) if precisa_escolher_curso else None

        # Validações
        erros = []
        if not foto:
            erros.append("É obrigatório enviar uma foto de rosto.")
        if not bi:
            erros.append("O número do B.I. é obrigatório.")
        if not email:
            erros.append("O e-mail é obrigatório.")
        if not aluno and not nome_visitante:
            erros.append("O nome completo é obrigatório.")
        if not aluno and not classe_atual:
            erros.append("Deves indicar a tua classe atual.")
        if precisa_escolher_curso and not curso_id:
            erros.append("Deves escolher o curso para a 10ª classe.")

        if erros:
            for e in erros:
                messages.error(request, e)
        else:
            curso_novo = None
            if precisa_escolher_curso and curso_id:
                curso_novo = get_object_or_404(Curso, id=curso_id)
            elif aluno and aluno.turma:
                curso_novo = aluno.turma.curso

            confirmacao = ConfirmacaoMatricula.objects.create(
                aluno=aluno,
                nome_completo=nome_visitante or (aluno.usuario.get_full_name() if aluno else None),
                ano_letivo=ano_letivo,
                foto_rosto=foto,
                bi_numero=bi,
                email=email,
                classe_nova=proxima_classe or classe_atual or '10',
                curso_novo=curso_novo,
                status='EM_REVISAO',
            )

            # Email automático — Status 1
            _email_confirmacao(confirmacao, 'EM_REVISAO')

            return render(request, 'academico/confirmacao_sucesso.html', {
                'confirmacao': confirmacao,
            })

    return render(request, 'academico/confirmacao_form.html', {
        'aluno': aluno,
        'usuario_logado': usuario_logado,
        'classe_atual': classe_atual,
        'proxima_classe': proxima_classe,
        'precisa_escolher_curso': precisa_escolher_curso,
        'cursos_ii_ciclo': cursos_ii_ciclo,
        'ano_letivo': ano_letivo,
        'classes_disponiveis': classes_disponiveis,
        'form_data': form_data,
    })


@login_required
def confirmacao_estado(request):
    """Página de estado da confirmação de matrícula do aluno."""
    try:
        aluno = Aluno.objects.get(usuario=request.user)
    except Aluno.DoesNotExist:
        return redirect('dashboard')

    confirmacao = ConfirmacaoMatricula.objects.filter(
        aluno=aluno
    ).order_by('-ano_letivo').first()

    return render(request, 'academico/confirmacao_estado.html', {
        'confirmacao': confirmacao,
    })


# ── Feature 2: Acesso Congelado ──────────────────────────────────────────

def acesso_congelado(request):
    """Página exibida quando o aluno tem matrícula congelada."""
    return render(request, 'academico/acesso_congelado.html')


# ── Feature 4: Exportação Excel e PDF ────────────────────────────────────

@login_required
def exportar_relatorio_excel(request):
    """Exporta os dados do relatório admin para Excel (.xlsx)."""
    if not (request.user.is_staff or request.user.is_admin_escola):
        raise PermissionDenied

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl não instalado. Corre: pip install openpyxl", status=500)

    ano = int(request.GET.get('ano', timezone.now().year))

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E8621A", end_color="E8621A", fill_type="solid")
    center = Alignment(horizontal='center')

    def _header_row(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    # ── Aba 1: Inscrições ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Inscrições"
    _header_row(ws1, ['Nome', 'B.I.', 'Curso', 'Classe', 'Status', 'Data Submissão'])
    for ins in Inscricao.objects.select_related('curso_pretendido').order_by('-data_submissao'):
        ws1.append([
            ins.nome_completo,
            ins.bi_numero,
            str(ins.curso_pretendido),
            ins.classe_pretendida,
            ins.get_status_display(),
            ins.data_submissao.strftime('%d/%m/%Y %H:%M'),
        ])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 20

    # ── Aba 2: Alunos por Turma ───────────────────────────────────────────
    ws2 = wb.create_sheet("Alunos por Turma")
    _header_row(ws2, ['Turma', 'Curso', 'Classe', 'Nº Processo', 'Nome', 'Email'])
    for aluno in Aluno.objects.filter(
        turma__ano_letivo=ano
    ).select_related('turma__curso', 'usuario').order_by('turma__nome', 'usuario__first_name'):
        ws2.append([
            str(aluno.turma) if aluno.turma else '-',
            str(aluno.turma.curso) if aluno.turma else '-',
            aluno.turma.classe if aluno.turma else '-',
            aluno.numero_processo,
            aluno.usuario.get_full_name(),
            aluno.usuario.email or '-',
        ])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    # ── Aba 3: Presenças ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Frequência")
    _header_row(ws3, ['Aluno', 'Disciplina', 'Data', 'Presente', 'Justificada'])
    for p in Presenca.objects.filter(
        aluno__turma__ano_letivo=ano
    ).select_related('aluno__usuario', 'disciplina').order_by('aluno__usuario__first_name'):
        ws3.append([
            p.aluno.usuario.get_full_name(),
            p.disciplina.nome,
            p.data.strftime('%d/%m/%Y'),
            'Sim' if p.esta_presente else 'Não',
            'Sim' if p.justificada else 'Não',
        ])
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    # ── Resposta ──────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Tarimba_{ano}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_relatorio_pdf(request):
    """Exporta o relatório admin para PDF."""
    if not (request.user.is_staff or request.user.is_admin_escola):
        raise PermissionDenied

    ano = int(request.GET.get('ano', timezone.now().year))

    # Reutiliza os dados do relatorio_admin
    from django.db.models import Count, Avg

    turmas_dados = Turma.objects.filter(ano_letivo=ano).annotate(
        num_alunos=Count('alunos')
    ).order_by('curso__nome', 'classe')

    inscricoes_por_status = Inscricao.objects.values('status').annotate(
        total=Count('id')
    )

    alunos_data = []
    for aluno in Aluno.objects.filter(
        turma__ano_letivo=ano
    ).select_related('turma', 'usuario').order_by('turma__nome', 'usuario__first_name'):
        notas = Nota.objects.filter(aluno=aluno)
        medias = [float(n.media_trimestral) for n in notas]
        media_geral = round(sum(medias) / len(medias), 1) if medias else 0
        faltas = Presenca.objects.filter(aluno=aluno, esta_presente=False).count()
        alunos_data.append({
            'nome': aluno.usuario.get_full_name(),
            'turma': str(aluno.turma) if aluno.turma else '-',
            'media': media_geral,
            'faltas': faltas,
        })

    context = {
        'ano': ano,
        'turmas': turmas_dados,
        'inscricoes_status': inscricoes_por_status,
        'alunos': alunos_data,
        'total_alunos': len(alunos_data),
        'data_geracao': timezone.now().strftime('%d/%m/%Y %H:%M'),
    }

    template = get_template('academico/relatorio_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Tarimba_{ano}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Erro ao gerar PDF: {pisa_status.err}', status=500)

    return response


@login_required
def exportar_relatorio_turma_excel(request, turma_id):
    """Exporta o relatório de uma turma para Excel (para o professor)."""
    if not request.user.is_professor:
        raise PermissionDenied

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl não instalado.", status=500)

    professor = get_object_or_404(Professor, usuario=request.user)
    turma = get_object_or_404(Turma, id=turma_id)

    if not Atribuicao.objects.filter(professor=professor, turma=turma).exists():
        raise PermissionDenied

    disciplinas = Disciplina.objects.filter(
        atribuicao__professor=professor, atribuicao__turma=turma
    )
    alunos = Aluno.objects.filter(turma=turma).select_related('usuario')

    wb = openpyxl.Workbook()
    ws = wb.active
    sheet_title = str(turma)
    if len(sheet_title) > 31:
        sheet_title = sheet_title[:28] + '...'
    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E8621A", end_color="E8621A", fill_type="solid")

    headers = ['Nº Processo', 'Nome'] + [d.nome for d in disciplinas] + ['Média Geral', 'Faltas']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for aluno in alunos:
        row = [aluno.numero_processo, aluno.usuario.get_full_name()]
        medias_aluno = []
        for disc in disciplinas:
            notas = Nota.objects.filter(aluno=aluno, disciplina=disc)
            if notas.exists():
                m = round(sum(float(n.media_trimestral) for n in notas) / notas.count(), 1)
                row.append(m)
                medias_aluno.append(m)
            else:
                row.append('-')

        media_geral = round(sum(medias_aluno) / len(medias_aluno), 1) if medias_aluno else 0
        faltas = Presenca.objects.filter(
            aluno=aluno, disciplina__in=disciplinas, esta_presente=False
        ).count()
        row += [media_geral, faltas]
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Relatorio_{turma.nome}_{turma.ano_letivo}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_relatorio_turma_pdf(request, turma_id):
    """Exporta o relatório de uma turma para PDF (para o professor)."""
    if not request.user.is_professor:
        raise PermissionDenied

    professor = get_object_or_404(Professor, usuario=request.user)
    turma = get_object_or_404(Turma, id=turma_id)

    if not Atribuicao.objects.filter(professor=professor, turma=turma).exists():
        raise PermissionDenied

    disciplinas = Disciplina.objects.filter(
        atribuicao__professor=professor, atribuicao__turma=turma
    )
    alunos = Aluno.objects.filter(turma=turma).select_related('usuario')

    dados_alunos = []
    for aluno in alunos:
        notas = Nota.objects.filter(aluno=aluno, disciplina__in=disciplinas)
        medias = [float(n.media_trimestral) for n in notas]
        media_geral = round(sum(medias) / len(medias), 1) if medias else 0
        faltas = Presenca.objects.filter(
            aluno=aluno, disciplina__in=disciplinas, esta_presente=False
        ).count()
        dados_alunos.append({
            'nome': aluno.usuario.get_full_name(),
            'processo': aluno.numero_processo,
            'media': media_geral,
            'faltas': faltas,
            'aprovado': media_geral >= 10,
        })

    dados_alunos.sort(key=lambda x: x['media'], reverse=True)

    context = {
        'turma': turma,
        'professor': professor,
        'disciplinas': disciplinas,
        'dados_alunos': dados_alunos,
        'data_geracao': timezone.now().strftime('%d/%m/%Y %H:%M'),
    }

    template = get_template('academico/relatorio_turma_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_{turma.nome}_{turma.ano_letivo}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Erro ao gerar PDF: {pisa_status.err}', status=500)

    return response
